import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def parse_line(line: str):
    line = line.strip()
    if not line:
        return None

    # Format: "Name (EXCHANGE: SYMBOL)"  e.g. "Bloom Energy (NYSE: BE)"
    match = re.match(r"^(.+?)\s*\(([A-Z]+):\s*([A-Z0-9.]+)\)\s*$", line)
    if match:
        return {
            "name": match.group(1).strip(),
            "exchange": match.group(2).strip(),
            "symbol": match.group(3).strip(),
        }

    # Format: "(EXCHANGE: SYMBOL)"  e.g. "(NYSEARCA: XLE)"
    match = re.match(r"^\(([A-Z]+):\s*([A-Z0-9.]+)\)\s*$", line)
    if match:
        return {
            "name": "",
            "exchange": match.group(1).strip(),
            "symbol": match.group(2).strip(),
        }

    # Format: "EXCHANGE: SYMBOL"  e.g. "NYSE: STZ"
    match = re.match(r"^([A-Z]+):\s*([A-Z0-9.]+)\s*$", line)
    if match:
        return {
            "name": "",
            "exchange": match.group(1).strip(),
            "symbol": match.group(2).strip(),
        }

    return None


def create_excel(input_file: str, output_file: str) -> None:
    with open(input_file, "r") as f:
        lines = f.readlines()

    rows = [r for line in lines if (r := parse_line(line)) is not None]

    # Remove duplicates where all three columns match
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r["name"], r["exchange"], r["symbol"])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    rows = sorted(unique_rows, key=lambda r: r["symbol"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stocks"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    headers = ["Symbol", "Exchange", "Name"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["symbol"])
        ws.cell(row=row_idx, column=2, value=row["exchange"])
        ws.cell(row=row_idx, column=3, value=row["name"])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_file)
    print(f"Saved {len(rows)} rows to {output_file}")


if __name__ == "__main__":
    create_excel("stocks.txt", "stocks.xlsx")
